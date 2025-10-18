import csv
from openpyxl import load_workbook, Workbook

from src.models import AnswerResult


def save_to_csv(
        csv_filename: str,
        transcript: str,
        transcript_column: int,
        results: list[AnswerResult]
) -> None:
    """
    Saves results to a csv file according to save_to_column field

    :param csv_filename: Path .csv file
    :param transcript: The transcript text to save
    :param transcript_column: Column index (1-based) for transcript
    :param results: List of AnswerResult objects
    :return: None
    """
    # Add to the columns transcript text
    results.append(AnswerResult(answer=transcript, save_to_column=transcript_column))
    results.sort(key=lambda r: r.save_to_column)

    # Get columns count
    with open(csv_filename, "r", newline="") as csvfile:
        reader = csv.reader(csvfile)
        data = list(reader)  # Read all rows into a list
        if data:  # Check if the list is not empty
            last_row = data[-1]
            columns_count = len(last_row)
        else:
            # If no one rows in file, use maximum value of save_to_column as columns_count
            columns_count = max(results, key=lambda r: r.save_to_column).save_to_column

    cur_pointer = 0
    row = []

    # Add only values to row when it equals result save_to_column otherwise put empty string
    for i in range(columns_count):
        if i == results[cur_pointer].save_to_column-1:
            row.append(results[cur_pointer].answer)
            cur_pointer += 1
            continue

        row.append("")

    with open(csv_filename, "a", newline="") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(row)


def save_to_excel(
        excel_filename: str,
        transcript: str,
        transcript_column: int,
        results: list[AnswerResult]
) -> None:
    """
    Saves results to an Excel (.xlsx) file according to save_to_column field.

    :param excel_filename: Path to .xlsx file
    :param transcript: The transcript text to save
    :param transcript_column: Column index (1-based) for transcript
    :param results: List of AnswerResult objects
    :return: None
    """
    # Add transcript as one of the results
    results.append(AnswerResult(answer=transcript, save_to_column=transcript_column))
    results.sort(key=lambda r: r.save_to_column)

    # Try to open existing workbook, or create a new one if not found
    try:
        wb = load_workbook(excel_filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active

    # Determine number of columns - either from sheet or from max(save_to_column)
    columns_count = ws.max_column if ws.max_column > 0 else 0
    columns_count = max(
        columns_count,
        max(r.save_to_column for r in results)
    )

    # Prepare the row with all empty cells
    row = ["" for _ in range(columns_count)]

    # Fill in cells according to save_to_column
    for r in results:
        if 0 < r.save_to_column <= columns_count:
            row[r.save_to_column - 1] = r.answer

    # Append the row at the end
    ws.append(row)
    wb.save(excel_filename)


def save_to_file(
        filename: str,
        transcript: str,
        transcript_column: int,
        results: list[AnswerResult]
) -> None:
    """
    Saves results to a file.

    :param filename: Path to .csv of .xlsx file
    :param transcript: The transcript text to save
    :param transcript_column: Column index (1-based) for transcript
    :param results: List of AnswerResult objects
    :return: None

    :raise: TypeError
    """
    filetype = filename.split(".")[-1]
    if filetype == "csv":
        save_to_csv(filename, transcript, transcript_column, results)
    elif filetype == "xlsx":
        save_to_excel(filename, transcript, transcript_column, results)
    else:
        raise TypeError(f"Unsupported file type: {filetype}")
